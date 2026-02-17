"""
Excel 解析服务
负责读取 Excel 文件，遍历 sheet 和行，提供预览和格式化功能
"""

from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetInfo:
    """Sheet 信息"""
    name: str
    row_count: int
    column_count: int


@dataclass
class ExcelPreview:
    """Excel 预览数据"""
    sheets: List[SheetInfo]
    rows: List[List[Any]]
    sheet_name: str


class ExcelParserError(Exception):
    """Excel 解析错误"""
    pass


class ExcelParser:
    """Excel 解析服务"""

    # 连续空行阈值
    EMPTY_ROW_THRESHOLD = 10

    def __init__(self, file_path: str):
        """
        初始化 Excel 解析器

        Args:
            file_path: Excel 文件路径
        """
        self.file_path = Path(file_path)
        self._workbook: Optional[Workbook] = None

        if not self.file_path.exists():
            raise ExcelParserError(f"文件不存在: {file_path}")

        if self.file_path.suffix.lower() not in [".xlsx", ".xls"]:
            raise ExcelParserError(f"不支持的文件格式: {self.file_path.suffix}")

    def open(self) -> None:
        """打开 Excel 文件"""
        try:
            self._workbook = openpyxl.load_workbook(
                str(self.file_path),
                read_only=True,
                data_only=True
            )
        except Exception as e:
            raise ExcelParserError(f"打开 Excel 文件失败: {str(e)}")

    def close(self) -> None:
        """关闭 Excel 文件"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False

    @property
    def workbook(self) -> Workbook:
        """获取 Workbook 对象"""
        if not self._workbook:
            self.open()
        return self._workbook

    def get_sheets(self) -> List[str]:
        """
        获取所有 Sheet 名称

        Returns:
            Sheet 名称列表
        """
        return self.workbook.sheetnames

    def get_sheet_info(self) -> List[SheetInfo]:
        """
        获取所有 Sheet 的信息

        Returns:
            SheetInfo 列表
        """
        infos = []
        for name in self.get_sheets():
            sheet = self.workbook[name]
            infos.append(SheetInfo(
                name=name,
                row_count=sheet.max_row,
                column_count=sheet.max_column
            ))
        return infos

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        获取指定的 Sheet

        Args:
            sheet_name: Sheet 名称

        Returns:
            Worksheet 对象
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ExcelParserError(f"Sheet 不存在: {sheet_name}")
        return self.workbook[sheet_name]

    def read_rows(
        self,
        sheet: Worksheet,
        start_row: int = 1,
        count: Optional[int] = None
    ) -> List[List[Any]]:
        """
        读取 Sheet 中的行数据

        Args:
            sheet: Worksheet 对象
            start_row: 起始行号（从 1 开始）
            count: 读取行数，None 表示读取到末尾

        Returns:
            行数据列表，每行是一个单元格值列表
        """
        rows = []
        end_row = start_row + count - 1 if count else sheet.max_row

        for row_num in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                row_data.append(cell.value)
            rows.append(row_data)

        return rows

    def read_column(
        self,
        sheet: Worksheet,
        column_index: int,
        start_row: int = 1,
        end_row: Optional[int] = None
    ) -> List[Any]:
        """
        读取整列数据

        Args:
            sheet: Worksheet 对象
            column_index: 列索引（0-based）
            start_row: 起始行号（从 1 开始）
            end_row: 结束行号（None 表示读取到末尾）

        Returns:
            列数据列表
        """
        column_data = []
        actual_end_row = end_row or sheet.max_row

        # 列索引转换为 Excel 列号（1-based）
        excel_col = column_index + 1

        for row_num in range(start_row, min(actual_end_row + 1, sheet.max_row + 1)):
            cell = sheet.cell(row=row_num, column=excel_col)
            column_data.append(cell.value)

        return column_data

    def read_row_by_columns(
        self,
        sheet: Worksheet,
        row_num: int,
        column_indices: List[int]
    ) -> Dict[int, Any]:
        """
        按指定列索引读取行数据

        Args:
            sheet: Worksheet 对象
            row_num: 行号（从 1 开始）
            column_indices: 列索引列表（0-based）

        Returns:
            {列索引: 值} 的字典
        """
        row_data = {}

        for col_idx in column_indices:
            # 列索引转换为 Excel 列号（1-based）
            excel_col = col_idx + 1
            if excel_col <= sheet.max_column:
                cell = sheet.cell(row=row_num, column=excel_col)
                row_data[col_idx] = cell.value
            else:
                row_data[col_idx] = None

        return row_data

    def read_row_full(
        self,
        sheet: Worksheet,
        row_num: int
    ) -> List[Any]:
        """
        读取整行数据

        Args:
            sheet: Worksheet 对象
            row_num: 行号（从 1 开始）

        Returns:
            行数据列表
        """
        row_data = []
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            row_data.append(cell.value)
        return row_data

    def get_preview_rows(
        self,
        sheet: Worksheet,
        count: int = 5
    ) -> List[List[Any]]:
        """
        获取预览行（前 N 行）

        Args:
            sheet: Worksheet 对象
            count: 预览行数

        Returns:
            预览行数据
        """
        return self.read_rows(sheet, start_row=1, count=count)

    def is_empty_row(self, row: List[Any]) -> bool:
        """
        检测是否为空行

        Args:
            row: 行数据

        Returns:
            是否为空行
        """
        for cell in row:
            if cell is not None and str(cell).strip():
                return False
        return True

    def format_row_for_ai(
        self,
        row: List[Any],
        headers: Optional[List[str]] = None
    ) -> str:
        """
        格式化行数据用于 AI 处理

        Args:
            row: 行数据
            headers: 表头列表（如果有）

        Returns:
            格式化的字符串
            - 有表头: "表头1:值1; 表头2:值2; ..."
            - 无表头: "值1 | 值2 | 值3 | ..."
        """
        if headers:
            # 有表头格式
            pairs = []
            for i, cell in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    value = str(cell) if cell is not None else ""
                    if value.strip():
                        pairs.append(f"{header}:{value}")
            return "; ".join(pairs)
        else:
            # 无表头格式
            values = []
            for cell in row:
                value = str(cell) if cell is not None else ""
                if value.strip():
                    values.append(value)
            return " | ".join(values)

    def iterate_rows(
        self,
        sheet: Worksheet,
        start_row: int = 1,
        skip_empty: bool = True
    ):
        """
        迭代 Sheet 中的行

        Args:
            sheet: Worksheet 对象
            start_row: 起始行号
            skip_empty: 是否跳过空行

        Yields:
            (row_number, row_data) 元组
        """
        consecutive_empty = 0

        for row_num in range(start_row, sheet.max_row + 1):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                row_data.append(cell.value)

            # 检查空行
            if self.is_empty_row(row_data):
                consecutive_empty += 1
                if consecutive_empty >= self.EMPTY_ROW_THRESHOLD:
                    # 连续 N 个空行，停止处理当前 Sheet
                    break
                if skip_empty:
                    continue
            else:
                consecutive_empty = 0

            yield (row_num, row_data)

    def get_total_rows(self, sheet: Worksheet, start_row: int = 1) -> int:
        """
        获取有效行总数（排除尾部连续空行）

        Args:
            sheet: Worksheet 对象
            start_row: 起始行号

        Returns:
            有效行数
        """
        count = 0
        consecutive_empty = 0

        for row_num in range(start_row, sheet.max_row + 1):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                row_data.append(cell.value)

            if self.is_empty_row(row_data):
                consecutive_empty += 1
                if consecutive_empty >= self.EMPTY_ROW_THRESHOLD:
                    break
            else:
                consecutive_empty = 0
                count += 1

        return count


def get_excel_preview(file_path: str, sheet_name: Optional[str] = None) -> ExcelPreview:
    """
    获取 Excel 文件预览

    Args:
        file_path: Excel 文件路径
        sheet_name: 指定 Sheet 名称（默认第一个）

    Returns:
        ExcelPreview 对象
    """
    with ExcelParser(file_path) as parser:
        sheets = parser.get_sheet_info()

        # 默认取第一个 sheet
        target_sheet = sheet_name or sheets[0].name if sheets else None
        if not target_sheet:
            raise ExcelParserError("Excel 文件没有 Sheet")

        sheet = parser.get_sheet(target_sheet)
        rows = parser.get_preview_rows(sheet, count=10)

        return ExcelPreview(
            sheets=sheets,
            rows=rows,
            sheet_name=target_sheet
        )
