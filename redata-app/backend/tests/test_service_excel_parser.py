"""
Excel 解析器测试
测试 ExcelParser 服务的解析功能
"""

import pytest
import os
import tempfile


class TestExcelParser:
    """Excel 解析器测试类"""

    @pytest.fixture
    def sample_xlsx_path(self):
        """创建示例 Excel 文件"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        # 创建临时文件
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 写入表头
        headers = ["姓名", "电话", "邮箱", "地址"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        data = [
            ["张三", "13812345678", "zhang@example.com", "北京市"],
            ["李四", "13987654321", "li@example.com", "上海市"],
            ["王五", "15012345678", "wang@example.com", "广州市"],
        ]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(path)
        yield path

        # 清理
        os.unlink(path)

    @pytest.fixture
    def multi_sheet_xlsx(self):
        """创建多 Sheet Excel 文件"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        wb = openpyxl.Workbook()

        # 第一个 Sheet
        ws1 = wb.active
        ws1.title = "数据表1"
        ws1.cell(row=1, column=1, value="字段A")
        ws1.cell(row=2, column=1, value="值1")

        # 第二个 Sheet
        ws2 = wb.create_sheet("数据表2")
        ws2.cell(row=1, column=1, value="字段B")
        ws2.cell(row=2, column=1, value="值2")

        wb.save(path)
        yield path
        os.unlink(path)

    # ========== 文件打开测试 ==========

    def test_open_file(self, sample_xlsx_path):
        """测试打开 Excel 文件"""
        from src.redata.services.excel_parser import ExcelParser

        parser = ExcelParser(sample_xlsx_path)
        parser.open()
        assert parser._workbook is not None
        parser.close()

    def test_open_file_not_found(self):
        """测试打开不存在的文件"""
        from src.redata.services.excel_parser import ExcelParser, ExcelParserError

        with pytest.raises(ExcelParserError):
            ExcelParser("/nonexistent/file.xlsx")

    def test_open_file_invalid_format(self):
        """测试打开无效文件格式"""
        from src.redata.services.excel_parser import ExcelParser, ExcelParserError

        # 创建一个非 Excel 文件
        fd, path = tempfile.mkstemp(suffix=".txt")
        os.write(fd, b"not an excel file")
        os.close(fd)

        try:
            with pytest.raises(ExcelParserError):
                ExcelParser(path)
        finally:
            os.unlink(path)

    def test_context_manager(self, sample_xlsx_path):
        """测试上下文管理器"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            assert parser._workbook is not None

    # ========== Sheet 信息测试 ==========

    def test_get_sheets(self, sample_xlsx_path):
        """测试获取 Sheet 名称列表"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheets = parser.get_sheets()
            assert len(sheets) >= 1
            assert "Sheet1" in sheets

    def test_get_sheet_info(self, sample_xlsx_path):
        """测试获取 Sheet 信息"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            info_list = parser.get_sheet_info()
            assert len(info_list) >= 1
            # 检查第一个 sheet 的信息
            info = info_list[0]
            assert info.name == "Sheet1"
            assert info.row_count >= 1
            assert info.column_count >= 1

    def test_get_sheet(self, sample_xlsx_path):
        """测试获取指定 Sheet"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            assert sheet is not None
            assert sheet.title == "Sheet1"

    # ========== 数据读取测试 ==========

    def test_read_rows(self, sample_xlsx_path):
        """测试读取多行"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            rows = parser.read_rows(sheet, start_row=1, count=2)
            assert len(rows) == 2
            assert "姓名" in rows[0]

    def test_read_row_full(self, sample_xlsx_path):
        """测试读取整行"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            row = parser.read_row_full(sheet, row_num=1)
            assert row is not None
            assert "姓名" in row

    def test_read_row_by_columns(self, sample_xlsx_path):
        """测试按列索引读取行"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            row_data = parser.read_row_by_columns(sheet, row_num=2, column_indices=[0, 1])
            assert 0 in row_data
            assert 1 in row_data
            assert row_data[0] == "张三"
            assert row_data[1] == "13812345678"

    def test_read_column(self, sample_xlsx_path):
        """测试读取整列"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            col = parser.read_column(sheet, column_index=0, start_row=1, end_row=4)
            assert len(col) == 4
            assert col[0] == "姓名"
            assert col[1] == "张三"

    def test_get_preview_rows(self, sample_xlsx_path):
        """测试获取预览行"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            preview = parser.get_preview_rows(sheet, count=3)
            assert len(preview) <= 3

    # ========== 多 Sheet 测试 ==========

    def test_multi_sheet_names(self, multi_sheet_xlsx):
        """测试多 Sheet 文件的 Sheet 名称"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(multi_sheet_xlsx) as parser:
            sheets = parser.get_sheets()
            assert len(sheets) == 2
            assert "数据表1" in sheets
            assert "数据表2" in sheets

    def test_read_from_different_sheets(self, multi_sheet_xlsx):
        """测试从不同 Sheet 读取数据"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(multi_sheet_xlsx) as parser:
            # 从第一个 Sheet 读取
            sheet1 = parser.get_sheet("数据表1")
            row1 = parser.read_row_full(sheet1, row_num=2)
            assert "值1" in row1

            # 从第二个 Sheet 读取
            sheet2 = parser.get_sheet("数据表2")
            row2 = parser.read_row_full(sheet2, row_num=2)
            assert "值2" in row2

    # ========== 空行检测测试 ==========

    def test_is_empty_row(self, sample_xlsx_path):
        """测试空行检测"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            # 空行
            assert parser.is_empty_row([None, None, "", ""]) == True
            assert parser.is_empty_row([]) == True

            # 非空行
            assert parser.is_empty_row(["值", None, "", None]) == False
            assert parser.is_empty_row(["", "值"]) == False

    # ========== 迭代测试 ==========

    def test_iterate_rows(self, sample_xlsx_path):
        """测试遍历数据行"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            rows = list(parser.iterate_rows(sheet, start_row=2))
            assert len(rows) >= 3

    def test_get_total_rows(self, sample_xlsx_path):
        """测试获取总行数"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            total = parser.get_total_rows(sheet)
            assert total >= 4  # 表头 + 3 行数据

    # ========== 格式化测试 ==========

    def test_format_row_for_ai_with_headers(self, sample_xlsx_path):
        """测试格式化行数据（有表头）"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            row = parser.read_row_full(sheet, row_num=2)
            headers = parser.read_row_full(sheet, row_num=1)
            formatted = parser.format_row_for_ai(row, headers)
            assert "姓名" in formatted
            assert "张三" in formatted

    def test_format_row_for_ai_without_headers(self, sample_xlsx_path):
        """测试格式化行数据（无表头）"""
        from src.redata.services.excel_parser import ExcelParser

        with ExcelParser(sample_xlsx_path) as parser:
            sheet = parser.get_sheet("Sheet1")
            row = parser.read_row_full(sheet, row_num=2)
            formatted = parser.format_row_for_ai(row, headers=None)
            assert "张三" in formatted


class TestExcelPreview:
    """Excel 预览功能测试"""

    def test_get_excel_preview(self):
        """测试获取 Excel 预览"""
        import openpyxl
        from src.redata.services.excel_parser import get_excel_preview

        # 创建临时文件
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试表"
        ws.cell(row=1, column=1, value="列1")
        ws.cell(row=2, column=1, value="值1")
        wb.save(path)

        try:
            preview = get_excel_preview(path)
            assert preview is not None
            assert preview.sheet_name == "测试表"
            assert len(preview.sheets) >= 1
            assert len(preview.rows) >= 1
        finally:
            os.unlink(path)
